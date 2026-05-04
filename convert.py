from pathlib import Path
import openpyxl
import xlrd
import fitz  # pymupdf


def convert_pdf(filepath: Path) -> str:
    doc = fitz.open(filepath)
    lines = []
    for page in doc:
        lines.append(page.get_text())
    return "\n\n".join(lines)


def convert_docx(filepath: Path) -> str:
    from docx import Document
    doc = Document(filepath)
    parts = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = para.style.name.lower()
        if "heading 1" in style:
            parts.append(f"# {text}")
        elif "heading 2" in style:
            parts.append(f"## {text}")
        elif "heading 3" in style:
            parts.append(f"### {text}")
        else:
            parts.append(text)
    return "\n\n".join(parts)


def convert_txt(filepath: Path) -> str:
    return filepath.read_text(encoding="utf-8", errors="replace")


def convert_xls(filepath: Path) -> str:
    wb = xlrd.open_workbook(filepath)
    parts = []
    for sheet in wb.sheets():
        parts.append(f"## {sheet.name}\n")
        rows = []
        for row_idx in range(sheet.nrows):
            row = [str(sheet.cell_value(row_idx, col)).strip()
                   for col in range(sheet.ncols)]
            rows.append(row)
        if not rows:
            continue
        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            while len(row) < len(header):
                row.append("")
            parts.append("| " + " | ".join(row) + " |")
        parts.append("")
    return "\n".join(parts)


def convert_xlsx(filepath: Path) -> str:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## {sheet_name}\n")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        max_cols = max(len(r) for r in rows)
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            while len(cells) < max_cols:
                cells.append("")
            parts.append("| " + " | ".join(cells) + " |")
        parts.append("")
    return "\n".join(parts)


CONVERTERS = {
    ".pdf":  convert_pdf,
    ".docx": convert_docx,
    ".txt":  convert_txt,
    ".xls":  convert_xls,
    ".xlsx": convert_xlsx,
}


def convert_file(filepath: Path) -> str:
    ext = filepath.suffix.lower()
    converter = CONVERTERS.get(ext)
    if not converter:
        raise ValueError(f"Formato não suportado: {ext}")
    return converter(filepath)